import openpyxl
import datetime
import os

from colorama import init
from colorama import Fore, Back, Style

from tqdm import tqdm

init()


def choose_file():

	def choose_one_file():
		while True:
			input_data = input("Какой файл необходимо обработать: ")
			if not input_data.isnumeric():
				print("Вы ввели не число. Попробуйте снова")
			elif int(input_data) in list_files:
				return input_data
				break
			elif int(input_data) == 0:
				print(Fore.WHITE)
				print(Back.BLACK)
				os.abort()
			else:
				print("Ваше число в неправильном диапазоне. Попробуйте снова")


	print(Fore.BLACK)
	print(Back.YELLOW)

	list_files = {}

	i = 1

	print(' 0 - Выход ')
	for root, dirs, files in os.walk("."):
		for filename in files:
			if root == '.':
				list_files[i] = filename
				print(' ' +str(i) + ' - ' +filename)
			i += 1

	print(Fore.WHITE)
	print(Back.GREEN)

	key_choose_file = choose_one_file()
	return list_files[int(key_choose_file)]




def menu():
	print(Fore.BLACK)
	print(Back.YELLOW)

	print(' 1 - Выбор файла для обработки')
	print(' 0 - Выход ')

	print(Fore.WHITE)
	print(Back.GREEN)

	q = input(' Выбери действие ? ')

	if q == '1':

		print(Fore.BLACK)
		print(Back.YELLOW)
		
		
		filename = choose_file()		

		delete_word = input('Введи слово : ')
		delete_word = delete_word.lower().strip()

		#открываем файл
		workbook = openpyxl.load_workbook(filename)
		# получить доступ к отдельному листу
		first_sheet = workbook.worksheets[0]

		delete_list = []
		data_list = []


		# получаем все листы и находим в них нужное название
		sheets = workbook.sheetnames
		for sheet in sheets:
			if sheet == 'Удалённые запросы':
				second_sheet = workbook['Удалённые запросы']
				for row in second_sheet.rows:
					word = row[0].value
					word = str(word)
					delete_list.append(word)


		# заполняем данными тематические списки
		for row in first_sheet.rows:
			word = row[0].value
			word = str(word)

			if word == 'None':
				continue
			else :

				data_word = word.strip().split(' ')

				len_count = len(delete_list)

				for single_word in data_word:

					if single_word == delete_word:
						delete_list.append(word)

				if len_count == len(delete_list):
					data_list.append(word)
					




		print(len(delete_list))
		print(len(data_list))







		# # создаю новую книгу
		workbook = openpyxl.Workbook()

		# # выбираем активный лист и меняем ему название
		ws_1 = workbook.active
		ws_1.title = "Запросы"

		# Заполняем основную таблицу
		
		i = 1;
		for word in data_list:
		    cellref = ws_1.cell(row=i, column=1)
		    cellref.value = word
		    i = i + 1


		ws_2 = workbook.create_sheet('Удалённые запросы', 1)
		
		i = 1;
		for word in delete_list:
		    cellref = ws_2.cell(row=i, column=1)
		    cellref.value = word
		    i = i + 1



		workbook.save(filename = filename)


	elif q == '0':
		pass
	else:
		menu()



menu()